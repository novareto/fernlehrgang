# -*- coding: utf-8 -*-
# Copyright (c) 2007-2010 NovaReto GmbH
# cklinger@novareto.de 

import grokcore.component as grok

from fernlehrgang.interfaces.flg import IFernlehrgang
from fernlehrgang.interfaces.resultate import ICalculateResults
from xlwt import Workbook
from zope.interface import Interface
from zope.schema import *
from zope.schema.interfaces import IVocabularyFactory, IContextSourceBinder
from zope.schema.vocabulary import SimpleVocabulary, SimpleTerm
from z3c.saconfig import Session
from fernlehrgang import models
from fernlehrgang import log
from sqlalchemy import and_


class IXLSReport(Interface):
        """ xls report """


@grok.provider(IContextSourceBinder)
def lehrhefte(context):
    rc = []
    for lehrheft in context.lehrhefte:
        titel = "%s, %s" %(lehrheft.nummer, lehrheft.titel)
        key = "%s-%s" %(lehrheft.id, lehrheft.nummer)
        rc.append(SimpleTerm(key, key, titel))
    return SimpleVocabulary(rc)


class IXLSExport(Interface):
    """ xml export """

    dateiname = TextLine(
        title=u"Dateiname",
        description=u"Dateiname"
        )

    rdatum = TextLine(
        title=u"RDatum",
        description=u"RDatum"
        )

    stichtag = TextLine(
        title=u"Stichtag",
        description=u"Stichtag"
        )

    lehrheft = Choice(
        title=u"Lehrheft",
        description=u"Lehrheft",
        source = lehrhefte,
        )


    def createXLS():
        """ """



spalten = ('FLG_ID', 'TEILNEHMER_ID', 'LEHRHEFT_ID', 'VERSANDANSCHRIFT', 'PLZ', 
    'MITGLNRMIT', 'FIRMA', 'FIRMA2', 'ANREDE', 'TITEL', 'VORNAME', 'NAME',
    'STRASSE', 'WOHNORT', 'PASSWORT', 'BELIEFART', 'R_DATUM', 'RSENDUNG', 'PUNKTZAHL',
    'STICHTAG', 'LEHRHEFT', 'R_TITEL', 'R_VORNAME', 'R_NAME', 'L1_F_1', 'L1_F_2',
    'L1_F_3', 'L1_F_4', 'L1_F_5', 'L1_F_6', 'L1_F_7', 'L1_F_8', 'L1_F_9', 'L1_F_10',
    'L2_F_1', 'L2_F_2', 'L2_F_3', 'L2_F_4', 'L2_F_5', 'L2_F_6', 'L2_F_7', 'L2_F_8', 'L2_F_9',
    'L2_F_10', 'L3_F_1', 'L3_F_2', 'L3_F_3', 'L3_F_4', 'L3_F_5', 'L3_F_6', 'L3_F_7',
    'L3_F_8', 'L3_F_9', 'L3_F_10', 'L4_F_1', 'L4_F_2', 'L4_F_3', 'L4_F_4', 'L4_F_5',
    'L4_F_6', 'L4_F_7', 'L4_F_8', 'L4_F_9', 'L4_F_10','L5_F_1', 'L5_F_2', 'L5_F_3',
    'L5_F_4', 'L5_F_5', 'L5_F_6', 'L5_F_7', 'L5_F_8', 'L5_F_9', 'L5_F_10', 'L6_F_1',
    'L6_F_2', 'L6_F_3', 'L6_F_4', 'L6_F_5', 'L6_F_6', 'L6_F_7', 'L6_F_8', 'L6_F_9',
    'L6_F_10', 'L7_F_1', 'L7_F_2', 'L7_F_3', 'L7_F_4', 'L7_F_5', 'L7_F_6', 'L7_F_7',
    'L7_F_8', 'L7_F_9', 'L7_F_10', 'L8_F_1', 'L8_F_2', 'L8_F_3', 'L8_F_4', 'L8_F_5',
    'L8_F_6', 'L8_F_7', 'L8_F_8', 'L8_F_9', 'L8_F_10','L1_P', 'L2_P', 'L3_P',
    'L4_P', 'L5_P', 'L6_P', 'L7_P', 'L8_P', 'L9_P', 'L10_P', 'BDANZSDG', 'BDNR', 'BDGEWICHT',
    'BZANZSDG', 'BZANZBD', 'BDANFANG', 'BDENDE',
)



def nN(value):
    """ Not None"""
    if value == None:
        return ''
    return value

from profilehooks import profile, timecall


class XLSExport(grok.Adapter):
    """ XML Export"""
    grok.implements(IXLSExport)
    grok.context(IFernlehrgang)

    def __init__(self, context):
        self.context = context
        self.book = Workbook()
        self.adressen = self.book.add_sheet('FLG-XLS Export')
         
    def createSpalten(self):
        for i, spalte in enumerate(spalten):
            self.adressen.write(0, i, spalte)

    def versandanschrift(self, teilnehmer):
        if teilnehmer.strasse != None or teilnehmer.nr != None or teilnehmer.plz != None or teilnehmer.ort != None:
            return "JA"
        return ""

    def createRows(self, form):
        flg = self.context
        lh_id, lh_nr = form['lehrheft'].split('-')
        ii = 0 
        session = Session()
        FERNLEHRGANG_ID = flg.id
        result = session.query(models.Teilnehmer, models.Unternehmen, models.Kursteilnehmer).filter(
            and_(
                models.Kursteilnehmer.fernlehrgang_id == FERNLEHRGANG_ID,
                models.Kursteilnehmer.teilnehmer_id == models.Teilnehmer.id,
                models.Teilnehmer.unternehmen_mnr == models.Unternehmen.mnr)).order_by(models.Teilnehmer.id).all()
        #lehrhefte = session.query(models.Lehrheft).filter(models.Lehrheft.fernlehrgang_id == FERNLEHRGANG_ID).all()
        i=1
        for teilnehmer, unternehmen, ktn in result:
            if ktn.status in ('A1', 'A2'):
                cal_res = ICalculateResults(ktn)
                summary = cal_res.summary()
                row = self.adressen.row(ii+1)
                row.write(0, nN(flg.id))
                row.write(1, nN(teilnehmer.id))
                row.write(2, lh_id)
                row.write(3, self.versandanschrift(teilnehmer))
                row.write(4, nN(teilnehmer.plz or unternehmen.plz))
                row.write(5, nN(unternehmen.mnr))
                row.write(6, nN(unternehmen.name))
                row.write(7, nN(unternehmen.name2))
                row.write(8, nN(teilnehmer.anrede))
                row.write(9, nN(teilnehmer.titel))
                row.write(10, nN(teilnehmer.vorname))
                row.write(11, nN(teilnehmer.name))
                strasse = nN(teilnehmer.strasse) + ' ' + nN(teilnehmer.nr)
                if strasse == " ":
                    strasse = nN(unternehmen.str)
                else:
                    if teilnehmer.adresszusatz:
                        strasse = strasse + ' // ' + teilnehmer.adresszusatz
                row.write(12, strasse)
                row.write(13, nN(teilnehmer.ort or unternehmen.ort))
                row.write(14, nN(teilnehmer.passwort))
                row.write(15, '') # Beliefart --> Leer laut Frau Esche 
                row.write(16, form['rdatum']) # Variable
                row.write(18, summary.get('resultpoints')) # PUNKTZAHL --> Punktzahl der Rücksendungen
                row.write(19, form['stichtag']) # Variable
                row.write(20, lh_nr) # LEHRHEFT --> Variable Für Welchen Ausdruck
                row.write(21, nN(teilnehmer.titel))
                row.write(22, nN(teilnehmer.vorname))
                row.write(23, nN(teilnehmer.name))
                z = 24 
                for lehrheft in flg.lehrhefte:
                    for frage in sorted(lehrheft.fragen, key=lambda frage: int(frage.frage)):
                        r=""
                        for antwort in ktn.antworten:
                            if frage.id == antwort.frage_id:
                                r = "%s\r %s\n %s\r" %(
                                        frage.antwortschema.upper(), 
                                        antwort.antwortschema, 
                                        cal_res.calculateResult(
                                            frage.antwortschema,
                                            antwort.antwortschema,
                                            frage.gewichtung))
                        row.write(z, r) 
                        z += 1
                lhid = ""        
                for lhr in cal_res.lehrhefte():
                    row.write(z, lhr.get('punkte'))
                    z += 1
                    if len(lhr['antworten']):
                       lhid = lhr['titel'].split('-')[0] 
                row.write(17, lhid) # RSENDUNG --> Anzahl der Rücksendung
                log('Fernlehrgang Anzahl', ii)
                ii+=1
            else:
                log('STATUS', ktn.status)


    def createXLS(self, form):
        self.createSpalten()
        self.createRows(form)
        file = open('/tmp/adr13-10.xls', 'w+')
        self.book.save(file)
        return file


from openpyxl.workbook import Workbook as NewWorkbook
from fernlehrgang.interfaces.kursteilnehmer import un_klasse, gespraech

v_un_klasse = un_klasse(None)
v_gespraech = gespraech(None)


class XLSReport(XLSExport):
    """ XML Export"""
    grok.implements(IXLSReport)
    grok.provides(IXLSReport)
    grok.context(IFernlehrgang)
    spalten = ['TEILNEHMER_ID', 'Titel', 'Anrede', 'Name', 'Vorname', 'Geburtsdatum', 'Strasse', 'Hausnummer', 'PLZ', 'ORT',
        'Mitgliedsnummer', 'Unternehmen', ' ', ' ', 'Strasse', 'PLZ', 'Ort', 'Registriert', 'Kategorie', 'Lieferstopps',
        'Mitarbeiteranzahl', 'Branche (Schrotthandel etc..)', u'Abschlussgespräch', 'Status', 'Punktzahl',
        u'Antwortbögen']

    def __init__(self, context):
        self.context = context
        self.book = NewWorkbook(optimized_write=True)
        self.adressen = self.book.create_sheet()
        self.rc = [self.spalten]

    def createSpalten(self):
        for i, spalte in enumerate(self.spalten):
            self.adressen.write(0, i, spalte)

    def ges_helper(self, term):
        try:
            return v_gespraech.getTerm(term).title
        except:
            pass
        return ''

    def un_helper(self, term):
        try:
            return v_un_klasse.getTerm(term).title
        except:
            pass
        return ''

    def createRows(self, form):
        flg = self.context
        session = Session()
        FERNLEHRGANG_ID = flg.id
        result = session.query(models.Teilnehmer, models.Unternehmen, models.Kursteilnehmer).filter(
            and_(
                models.Kursteilnehmer.fernlehrgang_id == FERNLEHRGANG_ID,
                models.Kursteilnehmer.teilnehmer_id == models.Teilnehmer.id,
                models.Teilnehmer.unternehmen_mnr == models.Unternehmen.mnr)).order_by(models.Teilnehmer.id).all()
        #lehrhefte = session.query(models.Lehrheft).filter(models.Lehrheft.fernlehrgang_id == FERNLEHRGANG_ID).all()
        i=1
        for teilnehmer, unternehmen, ktn in result:
            print i
            cal_res = ICalculateResults(ktn)
            summary = cal_res.summary()
            #import pdb; pdb.set_trace() 
            #row = self.adressen.row(i+2)
            liste = []
            teilnehmer = ktn.teilnehmer
            antworten = len(ktn.antworten)/10
            if teilnehmer:
                gebdat = ""
                if teilnehmer.geburtsdatum:
                    gebdat = teilnehmer.geburtsdatum.strftime('%d.%m.%Y')
                #unternehmen = teilnehmer.unternehmen
                liste.append(nN(teilnehmer.id))
                liste.append(nN(teilnehmer.titel))
                liste.append(nN(teilnehmer.anrede))
                liste.append(nN(teilnehmer.name))
                liste.append(nN(teilnehmer.vorname))
                liste.append(gebdat)
                liste.append(nN(teilnehmer.strasse))
                liste.append(nN(teilnehmer.nr))
                liste.append(nN(teilnehmer.plz))
                liste.append(nN(teilnehmer.ort))
                liste.append(nN(unternehmen.mnr))
                liste.append(nN(unternehmen.name))
                liste.append(nN(unternehmen.name2))
                liste.append(nN(unternehmen.name3))
                liste.append(nN(unternehmen.str))
                liste.append(nN(unternehmen.plz))
                liste.append(nN(unternehmen.ort))
                if teilnehmer.name:
                    liste.append('ja')
                else:
                    liste.append('nein')
                liste.append(nN(teilnehmer.kategorie))
                liste.append(ktn.status)
                liste.append(self.un_helper(ktn.un_klasse))
                liste.append(nN(ktn.branche))
                liste.append(self.ges_helper(ktn.gespraech))
                liste.append(nN(summary['comment']))
                liste.append(nN(summary['resultpoints']))
                liste.append(antworten)
            self.rc.append(liste)
            i+=1


    def createXLS(self, form):
        self.createRows(form)
        for z, line in enumerate(self.rc):
            self.adressen.append([cell for cell in line])
        file = '/tmp/report.xls'
        self.book.save(file)
        return open(file, 'r')


from zope.interface import Interface
import grok

class JEx(grok.View):
    grok.context(Interface)

    def render(self):
        session = Session()
        ID = "100026"
        FERNLEHRGANG_ID = "100"
        #result = session.query(models.Kursteilnehmer).join(models.Antwort).filter(
        result = session.query(models.Teilnehmer, models.Unternehmen, models.Kursteilnehmer).filter(
            and_(
                models.Kursteilnehmer.fernlehrgang_id == FERNLEHRGANG_ID,
                models.Kursteilnehmer.teilnehmer_id == models.Teilnehmer.id,
                models.Teilnehmer.unternehmen_mnr == models.Unternehmen.mnr)).order_by(models.Teilnehmer.id).all()
        for a,b,x in result:
            cal_res = ICalculateResults(x)
            print cal_res.summary()
            print x
            for y in x.antworten:
                print y
