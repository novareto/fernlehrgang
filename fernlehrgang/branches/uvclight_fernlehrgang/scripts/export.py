# -*- coding: utf-8 -*-
# Copyright (c) 2007-2010 NovaReto GmbH
# cklinger@novareto.de 


from z3c.saconfig import Session
from fernlehrgang import models
from sqlalchemy import and_
from fernlehrgang.interfaces.resultate import ICalculateResults
import csv
from openpyxl.workbook import Workbook


spalten = ('FLG_ID', 'TEILNEHMER_ID', 'LEHRHEFT_ID', 'VERSANDANSCHRIFT', 'PLZ',
    'MITGLNRMIT', 'FIRMA', 'FIRMA2', 'ANREDE', 'TITEL', 'VORNAME', 'NAME',
    'STRASSE', 'WOHNORT', 'PASSWORT', 'BELIEFART', 'R_DATUM', 'RSENDUNG', 'PUNKTZAHL',
    'STICHTAG', 'LEHRHEFT', 'R_TITEL', 'R_VORNAME', 'R_NAME', 'L1_F_1', 'L1_F_2',
    'L1_F_3', 'L1_F_4', 'L1_F_5', 'L1_F_6', 'L1_F_7', 'L1_F_8', 'L1_F_9', 'L1_F_10',
    'L2_F_1', 'L2_F_2', 'L2_F_3', 'L2_F_4', 'L2_F_5', 'L2_F_6', 'L2_F_7', 'L2_F_8', 'L2_F_9',
    'L2_F_10', 'L3_F_1', 'L3_F_2', 'L3_F_3', 'L3_F_4', 'L3_F_5', 'L3_F_6', 'L3_F_7',
    'L3_F_8', 'L3_F_9', 'L3_F_10', 'L4_F_1', 'L4_F_2', 'L4_F_3', 'L4_F_4', 'L4_F_5',
    'L4_F_6', 'L4_F_7', 'L4_F_8', 'L4_F_9', 'L4_F_10','L5_F_1', 'L5_F_2', 'L5_F_3',
    'L5_F_4', 'L5_F_5', 'L5_F_6', 'L5_F_7', 'L5_F_8', 'L5_F_9', 'L5_F_10', 'L6_F_1',
    'L6_F_2', 'L6_F_3', 'L6_F_4', 'L6_F_5', 'L6_F_6', 'L6_F_7', 'L6_F_8', 'L6_F_9',
    'L6_F_10', 'L7_F_1', 'L7_F_2', 'L7_F_3', 'L7_F_4', 'L7_F_5', 'L7_F_6', 'L7_F_7',
    'L7_F_8', 'L7_F_9', 'L7_F_10', 'L8_F_1', 'L8_F_2', 'L8_F_3', 'L8_F_4', 'L8_F_5',
    'L8_F_6', 'L8_F_7', 'L8_F_8', 'L8_F_9', 'L8_F_10','L1_P', 'L2_P', 'L3_P',
    'L4_P', 'L5_P', 'L6_P', 'L7_P', 'L8_P', 'L9_P', 'L10_P', 'BDANZSDG', 'BDNR', 'BDGEWICHT',
    'BZANZSDG', 'BZANZBD', 'BDANFANG', 'BDENDE',
)



#KONSTATNE
FERNLEHRGANG_ID = 100
LH_ID = 100
LH_NR = 1
RDATUM = "12.12.2002"
STICHTAG = "12.12.2004"

wb = Workbook(optimized_write = True)

#BOOK = Workbook()
#SHEET = BOOK.add_sheet('FLG-XLS Export')
#for i, spalte in enumerate(spalten):
#    SHEET.write(0, i, spalte)

#file = open('/tmp/a.csv', 'w')
#csv_file = csv.writer(file)



def nN(value):
    """ Not None"""
    if value == None:
        return ''
    return unicode(value).encode('utf-8')

def versandanschrift(teilnehmer):
    if teilnehmer.strasse != None or teilnehmer.nr != None or teilnehmer.plz != None or teilnehmer.ort != None:
         return "JA"
    return ""

from profilehooks import profile, timecall

rc=[spalten]
ws = wb.create_sheet()

#@timecall
@profile
def worker():
    i=1
    session = Session()
    result = session.query(models.Teilnehmer, models.Unternehmen, models.Kursteilnehmer).filter(
            and_(
                models.Kursteilnehmer.fernlehrgang_id == FERNLEHRGANG_ID,
                models.Kursteilnehmer.teilnehmer_id == models.Teilnehmer.id,
                models.Teilnehmer.unternehmen_mnr == models.Unternehmen.mnr)).order_by(models.Teilnehmer.id).all()
    lehrhefte = session.query(models.Lehrheft).filter(
            models.Lehrheft.fernlehrgang_id == FERNLEHRGANG_ID).all()
    for teilnehmer, unternehmen, ktn in result:
        liste = []
        cal_res = ICalculateResults(ktn)
        summary = cal_res.summary()

        liste.append(FERNLEHRGANG_ID)
        liste.append(nN(teilnehmer.id))
        liste.append(LH_ID)
        liste.append(versandanschrift(teilnehmer))
        liste.append(nN(teilnehmer.plz or unternehmen.plz))
        liste.append(nN(unternehmen.mnr))
        liste.append(nN(unternehmen.name))
        liste.append(nN(unternehmen.name2))
        liste.append(nN(teilnehmer.anrede))
        liste.append(nN(teilnehmer.titel))
        liste.append(nN(teilnehmer.vorname))
        liste.append(nN(teilnehmer.name))
        strasse = nN(teilnehmer.strasse) + ' ' + nN(teilnehmer.nr)
        if strasse == " ":
            strasse = nN(unternehmen.str)
        liste.append(strasse)
        liste.append(nN(teilnehmer.ort or unternehmen.ort))
        liste.append(nN(teilnehmer.passwort))
        liste.append('') # Beliefart --> Leer laut Frau Esche 
        liste.append(RDATUM) # Variable
        liste.append('PLATZHALTER')
        liste.append(summary.get('resultpoints')) # PUNKTZAHL --> Punktzahl der Rücksenn
        liste.append(STICHTAG) # Variable
        liste.append(LH_NR) # LEHRHEFT --> Variable Für Welchen Ausdruck
        liste.append(nN(teilnehmer.titel))
        liste.append(nN(teilnehmer.vorname))
        liste.append(nN(teilnehmer.name))
        for lehrheft in lehrhefte:
            for frage in sorted(lehrheft.fragen, key=lambda frage: int(frage.frage)):
                r = ""
                for antwort in ktn.antworten:
                    if frage.id == antwort.frage_id:
                        r = "%s\r %s\n %s\r" %(
                                frage.antwortschema.upper(),
                                antwort.antwortschema,
                                cal_res.calculateResult(
                                    frage.antwortschema,
                                    antwort.antwortschema,
                                    frage.gewichtung))
                liste.append(r)
        lhid = ""
        for lhr in cal_res.lehrhefte():
            liste.append(lhr.get('punkte'))
            if len(lhr['antworten']):
               lhid = lhr['titel'].split('-')[0]
        liste[17] = lhid 
        i += 1
        rc.append(liste)
    for z, line in enumerate(rc):
        ws.append([cell for cell in line])
    wb.save('/tmp/newck.xls')
    return rc


if __name__ == "__main__":
    print len(worker())
    exit()
